import pandas as pd
from os.path import exists
import threading

from pandas.io.common import Path, os
from Config import Config
from Log.Logger import Logger
from Types import Style 
from openpyxl import Workbook

exec_in_threads = False

class ExcelPrinter:
    settings = "_settings"
    headers = "_headers"
    content = "_content"
    main_headers = "_main_headers"
    logger: Logger
    page_name: str
    lock: threading.Lock = threading.Lock()
    lock_changed = False

    __process_col_pos = 0
    __process_row_pos = 0
    
    def __init__(self, layout: Config, page_name: str,logger: Logger = Logger("")):
        self.logger = logger
        self.layout = layout
        self.page_name = page_name

    def print(self, filepath, offset_x = 0, offset_y = 0, d_type: str ="layout"):
        try:
            data = self.layout.get_data(d_type)
        except:
            self.logger.print_critical_error(f"{filepath} does not contain usefull information")
            exit(1)
        self.__save_file(data, filepath, offset_x, offset_y)
        
    def __process_output_style(self, data: [], offset_y) -> dict[str: str]:
 
        for row in data:
            col_span = row[Style.COL_SPAN.value[0]]
            row_span = row[Style.ROW_SPAN.value[0]]
            break_line = row[Style.BREAK_LINE.value[0]]
            offset_col = row["offset-col"] 
            major = row[Style.MAJOR.value[0]]

            self.__add_style_process(row, col_span, row_span, not break_line, offset_col, major, offset_y)

        return data
    
    def set_lock(self, lock: threading.Lock):
        self.lock_changed = True
        self.lock = lock

    def execute_locked_process(self, callback) -> None:
        with self.lock:
            if not self.lock_changed or self.lock is None:
                self.logger.print_error("Lock not changed or not set. Please set it before calling this function.")
                return
            callback()

    def __get_property(self, row: dict, property: str, default_to):
        try:
            return row[property]
        except:
            return default_to

    same_row = False
    def __add_style_process(self, row: dict, col_span, row_span, same_row: bool = False, offset_col = 0, major = False, offset_row = 0):
        row.update({
            "col-start": self.__process_col_pos + 1 + offset_col,
            "col-end": self.__process_col_pos + col_span + offset_col,
            "row-start": self.__process_row_pos + 1 + offset_row,
            "row-end": self.__process_row_pos + row_span + offset_row,
        })

        if not same_row:
            self.__process_col_pos = 0
            self.__process_row_pos += row_span if not major else 1
        else:
            self.__process_col_pos += col_span + offset_col

        

    # função à prova de multi-threading
    # já implementa locks
    def __save_file(self, data, filepath, offset_x = 0, offset_y = 0):
        # esta função é executada mais abaixo com locks
        def __save():
            global exec_in_threads
            
            if not exec_in_threads:
                # verifica que o output ainda não existe
                # senão apaga o ficheiro
                if exists(filepath):
                    self.logger.print_info(f"{filepath} already exists, removing file...")
                    os.unlink(filepath)
                
                exec_in_threads = True


            from openpyxl import load_workbook

            # Primeira parte responsável por colocar e criar os dados
            # no arquivo excel
            
            num_cols = data[-1]["num-columns"]
            data_processed = self.process_data(data, num_cols)
            
            dict_data = {}

            random = 0
            for col in data_processed:
                try:
                    dict_data.update({random: col})
                    random += 1
                except:
                    continue

            # check if filepath exists
            if not exists(filepath):
                self.logger.print_info(f"Creating {filepath}...")
                Path(filepath).parent.mkdir(parents=True, exist_ok=True)
                wb = Workbook()
                wb.create_sheet(self.page_name)
                wb.save(filepath)
                wb.close()

            # Use 'with' statement for proper file handling

            #check if file is already open
            try:
               with open(filepath, "r") as _: # or just open
                   self.logger.print_info(f"{filepath} not open, all good")
            except IOError:
                self.logger.print_error(f"{filepath} is already open")

            with pd.ExcelWriter(filepath, mode='a', if_sheet_exists="overlay", engine="openpyxl") as writer:

                # remove arrays from the dict_data with len == 0
                keys_to_remove = []

                # Identify keys with empty arrays
                for key in dict_data:
                    if len(dict_data[key]) == 0:
                        keys_to_remove.append(key)

                # Remove keys with empty arrays
                for key in keys_to_remove:
                    del dict_data[key]

                try:
                    for key in dict_data:
                        self.logger.print_info("Column size: " + len(dict_data[key]).__str__())
                    pdDataframe = pd.DataFrame(dict_data)
                    pdDataframe.to_excel(writer, startrow = offset_y, startcol = offset_x, index=False, header=False, sheet_name=self.page_name)
                except Exception as e:
                    self.logger.print_critical_error(f"Error while writing to {filepath}: {e}")


            # Reopen the workbook to apply styles
            wb = load_workbook(filename=filepath) 
            try:
                ws = wb[self.page_name]
            except:
                # significa que a pagina não existe
                wb.create_sheet(self.page_name)
                ws = wb[self.page_name]

            self.__apply_style(ws, data[:-1], offset_x, offset_y)

            # Save and close the workbook
            wb.save(filepath)
            wb.close()
        

        self.execute_locked_process(callback= __save)



    def process_data(self, data, num_cols):
        cols = [[] for _ in range(num_cols)]
        index = 0

        for item in data:
            try:
                col_span = item["col-span"]
                row_span = item["row-span"]
                label = item["label"]
            except KeyError:
                continue
            except Exception:
                self.logger.print_critical_error("Item is not addressed correctly: ")
                print(item)
                exit(1)

            for i in range(col_span):
                for _ in range(row_span):
                    try:
                        cols[index + i].append(label)
                    except:
                        self.logger.print_critical_error(f"Erro: {index + i}, provavelmente existe algo de errado com o layout ou com o excel de input.\
                              De qualquer das formas peço já desculpa por não ter detetado este erro antes.")

            index += col_span

            # Supostamente, ver se há uma quebra de linha
            # é mais eficiente com item["break-line"]
            if index >= num_cols or item["break-line"]:
                index = 0

        return cols
    
    def __apply_style(self, ws, data:list[dict], offset_x = 0, offset_y = 0):
        from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

        major_data: dict[str: int] = {}
        style_list = []

        prev_item = None
        for style in data:

            if style[Style.MAJOR.value[0]] and style[Style.ID.value[0]] in major_data:
                item = major_data[style[Style.ID.value[0]]]

                if style[Style.BREAK_LINE.value[0]]:
                    if prev_item:
                        prev_item[Style.BREAK_LINE.value[0]] = True

                if item == 0:
                    del item

                item -= 1
                continue

            elif style[Style.MAJOR.value[0]]:
                major_data.update({
                    style[Style.ID.value[0]]: style[Style.MAJOR_SPAN.value[0]]
                })
                style["row-span"] = style[Style.MAJOR_SPAN.value[0]]

            style_list.append(style)
            prev_item = style

        self.__process_output_style(style_list, offset_y)

        for style in style_list:
            ws.merge_cells(start_row=style['row-start'], start_column=style['col-start'],
                        end_row=style['row-end'], end_column=style['col-end'])
            
            alignment = Alignment(horizontal=style['x-alignment'], vertical=style["y-alignment"])
            font = Font(color=style['text-color'])
            fill = PatternFill(patternType='solid', fill_type='solid', fgColor=style['bg-color'])
            border = Border(
                left=Side(style=style["border"], color=style["border-color"], border_style=None),   # Black thin border
                right=Side(style=style["border"], color=style["border-color"], border_style=None),
                top=Side(style=style["border"], color=style["border-color"], border_style=None),
                bottom=Side(style=style["border"], color=style["border-color"], border_style=None)
            )

            self.__apply_style_to_cell(ws, style['row-start'], style['col-start'], 'alignment', alignment)
            self.__apply_style_to_cell(ws, style['row-start'], style['col-start'], 'fill', fill)
            self.__apply_style_to_cell(ws, style['row-start'], style['col-start'], 'font', font)

            for i in range(style['col-start'], style['col-end'] + 1):
                self.__apply_style_to_cell(ws, style['row-start'], i, 'border', border)
            
            for j in range(style['row-start'], style['row-end'] + 1):
                self.__apply_style_to_cell(ws, j, style['col-start'], 'border', border)
            
    def __apply_style_to_cell(self, ws, row, col, stylename, style):
        cell = ws.cell(row=row, column=col)
        setattr(cell, stylename, style)

    
    def __apply_style_to_all_cells(self, ws, stylename, style):
        for row in ws.iter_rows():
            for cell in row:
                setattr(cell, stylename, style)


    
