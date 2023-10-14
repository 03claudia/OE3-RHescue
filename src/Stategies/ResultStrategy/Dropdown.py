from openpyxl.worksheet.datavalidation import DataValidation
from xlsxwriter.workbook import Worksheet
from Log.Logger import Logger
import xlsxwriter

class Dropdown:
    options: str 
    people: list[str]
    worksheet_name: str = "Avaliacao"

    def __init__(self, people: list[str], dropdown_pos: str):
        self.people = people
        self.options = "=CONCATENAR("
        self.dropdown_pos = dropdown_pos

    # objetivo
    # =CONCATENAR(
    #     SE(Avaliacao!A1="Catarina Milheiro", "6", ""),
    #     SE(Avaliacao!A1="Gonçalo Figueiredo", "5", ""),
    #     SE(Avaliacao!A1="Inês Cabral", "4", ""),
    #     SE(Avaliacao!A1="Mariana Arezes", "3", ""),
    #     SE(Avaliacao!A1="Mariana Oliveira", "2", ""),
    #     SE(Avaliacao!A1="Paula Ferreira", "1", ""),
    #     SE(Avaliacao!A1="Paulo Vieira", "7", ""),
    #     SE(Avaliacao!A1="Rafaela Carvalho", "8", "")
    # )
    def if_(self, dropdown_option: str, set_cell_to: str):
        self.options += f'SE({self.worksheet_name}!{self.dropdown_pos}="{dropdown_option}", {set_cell_to}, ""),'


    def get_options(self) -> str:
        # o codigo por aqui está demasiado feio
        return self.options + ')'

    def draw_dropdown(self, filepath):
        from openpyxl import load_workbook

        # Load the existing Excel workbook
        existing_workbook = load_workbook(filepath)
    
        # Get the worksheet you want to modify or create a new one if it doesn't exist
        worksheet_name = self.worksheet_name
        if worksheet_name not in existing_workbook.sheetnames:
            worksheet = existing_workbook.create_sheet(worksheet_name)
        else:
            worksheet = existing_workbook[worksheet_name]
   
    
        # Write the list of people to a column in the worksheet
        for i, person in enumerate(self.people):
            worksheet.cell(row=i+1, column=2, value=person)
   
        # Create a dropdown in cell A1 that references the list of people
        data_validation = DataValidation(
            type="list",
            formula1=f'=$B$1:$B${len(self.people)}',  # Assumes the list is in column A
        )
        worksheet.add_data_validation(data_validation)
        data_validation.add(worksheet["A1"])

        # Save the changes to the existing workbook
        existing_workbook.save(filepath)

    def reset(self):
        self.options = "=CONCATENAR("
