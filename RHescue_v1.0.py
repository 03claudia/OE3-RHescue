"FINAL PROVISÓRIO RHESCUE - TENTATIVA"

from tkinter import *
import tkinter as tk
from tkinter.ttk import *
import tkinter.font as font
from tkinter.filedialog import askopenfile
from tkinter import ttk, filedialog
import time
import os
import openpyxl as xl
from openpyxl import workbook
import pandas as pd
import openpyxl
import numpy as np
import copy as copy
import webbrowser

#######################################################################################################################################

"""
Parte 0: Definição da interface gráfica e de elementos base.
"""

# Criação da janela.
root = Tk()
root.title('RHescue')
root.iconbitmap("EPIC.ico")
root.geometry('1200x673')

# Adicionar imagem de fundo.
bg = PhotoImage(file="BG1.png")

# Disposição da imagem.
label1 = Label(root, image=bg)
label1.place(x=0, y=0)

# Definição da fonte da letra.
tipo_letra_normal = font.Font(family="Leelawadee", size=10)
tipo_letra_negrito = font.Font(family="Leelawadee", size=10, weight="bold")

# Definição da progress bar
s = Style()
# s.theme_use('clam')
s.configure("red.Horizontal.TProgressbar", foreground='white', background='orange')
progress = Progressbar(root, style="red.Horizontal.TProgressbar", orient="horizontal", mode="determinate", length=675)
progress.pack()
progress.place(x=200, y=600)
progress['value'] = 0

#######################################################################################################################################

"""
Parte 1: Definição de funções auxiliares -> Poderão ser utilizadas em mÃ³dulo separado.
"""


# Função que extrai nome de ficheiro perante input de caminho absoluto desse mesmo ficheiro no PC.
def nome_ficheiro(string):
    aux = []
    for i in range(1, len(string)):
        if string[-i] == "/" or string[-i] == "\\":
            break
        aux.append(string[-i])
    aux.reverse()
    final = ""
    for c in aux:
        final += c
    return final


# Função auxiliar para S2.1, S2.2, S2.3
def nome_em_frase(string):
    
    indice1=string.index("[")
    indice2=string.index("]")
    
    nome=""
    for carater in string[indice1+1:indice2]:
        nome+=carater
        
    return nome


# Função que obtÃ©m o caminho absoluto de um ficheiro escolhido.
def getFolderPath():
    global folder
    folder = filedialog.askdirectory()


def progress_step(step):
    progress['value'] += step
    Label(root, text=str(progress['value']) + '%', foreground='black').place(x=875, y=603)


#######################################################################################################################################

"""
Parte 2: Definição e disposição de widgets na interface gráfica
"""


# MENU DE ESCOLHA
def ShowChoice():
    global dep
    # print(v.get())
    dep = v.get()


# Definição de variáveis para "escolha de opção".
v = IntVar()
departamentos = [("Projetos", 1), ("Marketing", 2), ("Financeiro", 3), ("Recursos Humanos", 4)]

# Tí­tulo da choose option
Label(root, text="Departamento:", justify=LEFT, background="#FF914D", font=tipo_letra_negrito).place(x=35, y=350)
k = 0
for departamento, val in departamentos:
    k += 20
    Radiobutton(root, text=departamento, command=ShowChoice, variable=v, value=val).place(x=35, y=350 + k)


# BOTÃ•ES
###
def openfile_FINAL():
    # Variável GLOBAL -> Para utilização mais tarde.
    global filepath_FINAL

    # Codificação do botÃƒÂ£o.
    file = filedialog.askopenfile(mode='r', filetypes=[('Excel Files', '*xlsx')])
    if file:
        filepath_FINAL = os.path.abspath(file.name)

        # Adicionar Label de apoio
        Label(root, text=nome_ficheiro(str(filepath_FINAL))[:25], foreground='green', font=tipo_letra_normal).place(x=205,
                                                                                                               y=395)

    global caminho_FINAL
    caminho_FINAL = str(filepath_FINAL).replace("\\", "/")


FINAL = Label(root, text="Excel de Av. Semestral", background="#FF914D", font=tipo_letra_negrito).place(x=205, y=350)
FINALbtn = Button(root, text="Escolher ficheiro", command=lambda: openfile_FINAL()).place(x=205, y=370)


###
def openfile_ADM1():
    # Codificação do botão.
    file = filedialog.askopenfile(mode='r', filetypes=[('Excel Files', '*xlsx')])
    if file:
        filepath_ADM1 = os.path.abspath(file.name)

        # Adicionar Label de apoio
        Label(root, text=nome_ficheiro(str(filepath_ADM1))[:25], foreground='green', font=tipo_letra_normal).place(x=380,
                                                                                                              y=395)

    global caminho_ADM1
    caminho_ADM1 = str(filepath_ADM1).replace("\\", "/")


ADM1 = Label(root, text="M1: Desemp. Mensal", background="#FF914D", font=tipo_letra_negrito).place(x=380, y=350)
ADM1btn = Button(root, text="Escolher ficheiro", command=lambda: openfile_ADM1()).place(x=380, y=370)


###
def openfile_ADS1():
    # Variável GLOBAL -> Para utilização mais tarde.
    global filepath_ADS1

    # Codificação do botão.
    file = filedialog.askopenfile(mode='r', filetypes=[('Excel Files', '*xlsx')])
    if file:
        filepath_ADS1 = os.path.abspath(file.name)

        # Adicionar Label de apoio
        Label(root, text=nome_ficheiro(str(filepath_ADS1))[:25], foreground='green', font=tipo_letra_normal).place(x=380,
                                                                                                              y=470)

    global caminho_ADS1
    caminho_ADS1 = str(filepath_ADS1).replace("\\", "/")


ADS1 = Label(root, text="S1: Desemp. Semestral", background="#FF914D", font=tipo_letra_negrito).place(x=380, y=425)
ADS1btn = Button(root, text="Escolher ficheiro", command=lambda: openfile_ADS1()).place(x=380, y=445)


###
def openfile_S21():
    # Codificação do botão
    file = filedialog.askopenfile(mode='r', filetypes=[('Excel Files', '*xlsx')])
    if file:
        filepath_S21 = os.path.abspath(file.name)

        # Adicionar Label de apoio
        Label(root, text=nome_ficheiro(str(filepath_S21))[:25], foreground='green', font=tipo_letra_normal).place(x=555,
                                                                                                             y=395)

    global caminho_S21
    caminho_S21 = str(filepath_S21).replace("\\", "/")


S21 = Label(root, text="S2.1: Av. de Membros", background="#FF914D", font=tipo_letra_negrito).place(x=555, y=350)
S21btn = Button(root, text="Escolher ficheiro", command=lambda: openfile_S21()).place(x=555, y=370)


###
def openfile_S22():
    # Codificação do botão.
    file = filedialog.askopenfile(mode='r', filetypes=[('Excel Files', '*xlsx')])
    if file:
        filepath_S22 = os.path.abspath(file.name)

        # Adicionar Label de apoio
        Label(root, text=nome_ficheiro(str(filepath_S22))[:25], foreground='green', font=tipo_letra_normal).place(x=555,
                                                                                                             y=465)

    global caminho_S22
    caminho_S22 = str(filepath_S22).replace("\\", "/")


S22 = Label(root, text="S2.2: Av. de PM's", background="#FF914D", font=tipo_letra_negrito).place(x=555, y=425)
S22btn = Button(root, text="Escolher ficheiro", command=lambda: openfile_S22()).place(x=555, y=440)


###
def openfile_S23():
    # Codificação do botão.
    file = filedialog.askopenfile(mode='r', filetypes=[('Excel Files', '*xlsx')])
    if file:
        filepath_S23 = os.path.abspath(file.name)

        # Adicionar Label de apoio
        Label(root, text=nome_ficheiro(str(filepath_S23))[:25], foreground='green', font=tipo_letra_normal).place(x=555,
                                                                                                             y=540)

    global caminho_S23
    caminho_S23 = str(filepath_S23).replace("\\", "/")


S23 = Label(root, text="S2.3: Av. de Coord's", background="#FF914D", font=tipo_letra_negrito).place(x=555, y=495)
S23btn = Button(root, text="Escolher ficheiro", command=lambda: openfile_S23()).place(x=555, y=515)


###
def openfile_AM1():
    # Codificação do botão.
    file = filedialog.askopenfile(mode='r', filetypes=[('Excel Files', '*xlsx')])
    if file:
        filepath_AM1 = os.path.abspath(file.name)

        # Adicionar Label de apoio
        Label(root, text=nome_ficheiro(str(filepath_AM1))[:25], foreground='green', font=tipo_letra_normal).place(x=730,
                                                                                                             y=395)

    global caminho_AM1
    caminho_AM1 = str(filepath_AM1).replace("\\", "/")


AM1 = Label(root, text="AM1: Av. Motivacional", background="#FF914D", font=tipo_letra_negrito).place(x=730, y=350)
AM1btn = Button(root, text="Escolher ficheiro", command=lambda: openfile_AM1()).place(x=730, y=370)


###
def openfile_AM2():
    # Codificação do botão.
    file = filedialog.askopenfile(mode='r', filetypes=[('Excel Files', '*xlsx')])
    if file:
        filepath_AM2 = os.path.abspath(file.name)

        # Adicionar Label de apoio
        Label(root, text=nome_ficheiro(str(filepath_AM2))[:25], foreground='green', font=tipo_letra_normal).place(x=730,
                                                                                                             y=470)

    global caminho_AM2
    caminho_AM2 = str(filepath_AM2).replace("\\", "/")


AM2 = Label(root, text="AM2: Av. Motivacional", background="#FF914D", font=tipo_letra_negrito).place(x=730, y=425)
AM2btn = Button(root, text="Escolher ficheiro", command=lambda: openfile_AM2()).place(x=730, y=445)


###
def openfile_AM3():
    # Codificação do botão.
    file = filedialog.askopenfile(mode='r', filetypes=[('Excel Files', '*xlsx')])
    if file:
        filepath_AM3 = os.path.abspath(file.name)

        # Adicionar Label de apoio
        Label(root, text=nome_ficheiro(str(filepath_AM3))[:25], foreground='green', font=tipo_letra_normal).place(x=730,
                                                                                                             y=545)

    global caminho_AM3
    caminho_AM3 = str(filepath_AM3).replace("\\", "/")


AM3 = Label(root, text="AM3: Av. Motivacional", background="#FF914D", font=tipo_letra_negrito).place(x=730, y=500)
AM3btn = Button(root, text="Escolher ficheiro", command=lambda: openfile_AM3()).place(x=730, y=520)


###
def EXEC_fun(caminho_FINAL, caminho_ADM1, caminho_ADS1, caminho_AM1, caminho_AM2, caminho_AM3, caminho_S21, caminho_S22,
             caminho_S23, departamento):
    # Label de loading
    Label(root, text="Loading...", foreground='green', font=tipo_letra_negrito).place(x=530, y=575)
    progress_step(5)
    root.update_idletasks()

    #######################################################################################################################################
    """
    Parte 1: Inputs necessários
    """
    #######################################################################################################################################

    # Definição do departamento a efetuar a avaliação.
    # departamento=str(input("DEPARTAMENTO A AVALIAR"))

    # Lista de pm's de cada projeto -> prÃ© requisito para S2.1,S2.2 e S2.3
    base_dados_ex = pd.read_excel(caminho_FINAL, sheet_name="Projetos")
    lista_projeto_pm = base_dados_ex.values.tolist()
    #variáveis necessárias para os diretores
    diretor = []
    diretor_proj = []
    diretor_rh = []
    diretor_mkt = []
    diretor_fin = []
    # Diretores de cada departamento
    wb1 = xl.load_workbook(caminho_FINAL)
    ws1 = wb1.worksheets[0]
    x = 2
    while ws1.cell(row=x, column=6).value == None:
        x = x + 1
    while ws1.cell(row=x, column=6).value != None:
        if ws1.cell(row=x, column=6).value == 'PROJ':
            diretor_proj.append(ws1.cell(row=x, column=2).value)
        elif ws1.cell(row=x, column=6).value == 'RH':
            diretor_rh.append(ws1.cell(row=x, column=2).value)
        elif ws1.cell(row=x, column=6).value == 'MKT':
            diretor_mkt.append(ws1.cell(row=x, column=2).value)
        elif ws1.cell(row=x, column=6).value == 'FIN':
            diretor_fin.append(ws1.cell(row=x, column=2).value)
        x = x + 1

    # Diretor do departamento "atual"
    if dep == 1:
        diretor = diretor_proj.copy()
    elif dep == 2:
        diretor = diretor_mkt.copy()
    elif dep == 3:
        diretor = diretor_fin.copy()
    elif dep == 4:
        diretor = diretor_rh.copy()
    #######################################################################################################################################
    """
    Parte 2: Input e Escrita no ficheiro final dos exceis necessários.
    """
    #######################################################################################################################################
    """
    AM1 - Import de excel necessário e escrita no ficheiro final
    """

    # Input
    wb1 = xl.load_workbook(caminho_AM1)
    ws1 = wb1.worksheets[0]

    # Output
    wb2 = xl.load_workbook(caminho_FINAL)
    ws2 = wb2.active

    mr = ws1.max_row
    mc = ws1.max_column

    sheet = wb2["AM Mês" + str(1)]

    for y in range(1, mc + 1):
        for x in range(1, mr):
            c = ws1.cell(row=x, column=y + 1)
            sheet.cell(row=x, column=y).value = c.value
    wb2.save(caminho_FINAL)

    progress_step(11)
    root.update_idletasks()

    #######################################################################################################################################
    """
    AM2 - Import de excel necessário e escrita no ficheiro final
    """

    # Input
    wb1 = xl.load_workbook(caminho_AM2)
    ws1 = wb1.worksheets[0]

    # Output
    wb2 = xl.load_workbook(caminho_FINAL)
    ws2 = wb2.active

    mr = ws1.max_row
    mc = ws1.max_column

    sheet = wb2["AM Mês" + str(2)]

    for y in range(1, mc + 1):
        for x in range(1, mr):
            c = ws1.cell(row=x, column=y + 1)
            sheet.cell(row=x, column=y).value = c.value
    wb2.save(caminho_FINAL)

    progress_step(11)
    root.update_idletasks()

    #######################################################################################################################################
    """
    AM3 - Import de excel necessário e escrita no ficheiro final
    """

    # Input
    wb1 = xl.load_workbook(caminho_AM3)
    ws1 = wb1.worksheets[0]

    # Output
    wb2 = xl.load_workbook(caminho_FINAL)
    ws2 = wb2.active

    mr = ws1.max_row
    mc = ws1.max_column

    sheet = wb2["AM Mês" + str(3)]

    for y in range(1, mc + 1):
        for x in range(1, mr):
            c = ws1.cell(row=x, column=y + 1)
            sheet.cell(row=x, column=y).value = c.value
    wb2.save(caminho_FINAL)

    progress_step(12)
    root.update_idletasks()

    #######################################################################################################################################
    """
    S2.1 - Import de excel necessário e escrita no ficheiro final
    """

    # Definição de funções

    def nome_em_frase(string):

        nome1 = ""
        k = 0
        for carater in string:
            if carater == "[":
                k = 1
                continue
            if k == 1:
                nome1 += carater
        nome2 = ""
        for carater in nome1:
            if carater != "]" and carater != ".":
                nome2 += carater

        return nome2

    # LEITURA DA BASE DE DADOS
    base_dados_ex = pd.read_excel(caminho_S21)
    base_dados_original = base_dados_ex.values.tolist()

    # Copiar base de dados
    base_dados = copy.deepcopy(base_dados_original)

    # Lista de colunas
    lista_colunas = base_dados_ex.columns
    lista_colunas = lista_colunas.tolist()
    del lista_colunas[0]

    # Limpar timestamps
    for linha in base_dados:
        linha.pop(0)

    # Limpar observações da base de dados a utilizar, mas criando uma lista com as observações.
    lista_observações = []
    for linha in base_dados:

        # Nome do projeto
        lista_aux = []
        lista_aux.append(linha[0])

        # Avaliador
        for i in range(1, len(linha)):
            try:
                if np.isnan(linha[i]) == True:
                    continue
            except:
                lista_aux.append(linha[i])
                break

        # Observações
        lista_aux.append(linha[-2])
        lista_aux.append(linha[-1])

        lista_observações.append(lista_aux)

    for linha in base_dados:
        del linha[-1]
        del linha[-1]

    # Projetos
    lista_projetos = []
    for linha in base_dados:
        if linha[0] not in lista_projetos:
            lista_projetos.append(linha[0])

    # PM's de cada projeto

    # lista_projeto_pm=[["Responsabilidade Social (Interno)","Pedro Machado"],["Site EPIC","Liliana Freitas"],["WEB4SOCIETY","Pedro Machado"],["RefractorAJ","Filipe AraÃƒÂºjo"]]
    # Membros de cada projeto

    lista_membros_projetos_raw = []  # Lista crua - Não trabalhada.
    for linha in base_dados:

        # Nome projeto
        aux = []
        aux.append(linha[0])

        # Membro
        for dado in linha[1:]:
            try:
                if np.isnan(dado) == True:
                    continue
            except:
                aux.append(dado)
                break

        lista_membros_projetos_raw.append(aux)

    lista_membros_projetos = []  # Lista modificada - Trabalhada.
    for projeto1 in lista_projetos:
        lista_aux = []
        membros_aux = []

        for projeto2, membro in lista_membros_projetos_raw:

            if projeto1 == projeto2:
                membros_aux.append(membro)

        lista_aux.append(projeto1)

        lista_aux1 = []
        for membro in membros_aux:
            lista_aux1.append(membro)

        lista_aux.append(lista_aux1)
        lista_membros_projetos.append(lista_aux)

    # Criação de uma lista que, por cada avaliador, indica o valor atribuido a par com o avaliado.
    lista_avaliações = []

    for projeto1 in lista_projetos:
        for projeto2, lista_membros_projeto in lista_membros_projetos:
            for membro in lista_membros_projeto:
                lista_aux = []
                if projeto1 == projeto2:  # temos indexados os membros e o projeto, podemos procurar os valores.

                    for linha in base_dados:
                        k = 0
                        if linha[0] == projeto1:
                            for i in range(len(linha)):
                                try:
                                    if np.isnan(linha[i]) == True:
                                        k = 0
                                        continue
                                except:
                                    if membro == linha[i]:
                                        lista_aux.append(projeto1)
                                        lista_aux.append(membro)
                                        k = 1
                                        continue
                                if k == 1:
                                    lista_aux.append([linha[i], nome_em_frase(lista_colunas[i])])

                lista_avaliações.append(lista_aux)

    k = 0
    for i in range(len(lista_avaliações)):
        if len(lista_avaliações[i + k]) == 0:
            del lista_avaliações[i + k]
            k -= 1

            # Adicionar as observações

    for i in range(len(lista_observações)):
        for j in range(len(lista_avaliações)):
            if lista_observações[i][0] == lista_avaliações[j][0] and lista_observações[i][1] == lista_avaliações[j][1]:
                lista_avaliações[j].append(lista_observações[i][-2])
                lista_avaliações[j].append(lista_observações[i][-1])

    # O prÃƒÂƒÃ‚Â³ximo passo ÃƒÂƒÃ‚Â© trabalhar a lista anterior de modo a deixar o output exatemente como o queremos.

    lista_final = []
    for avaliação in lista_avaliações:

        # Retirar a lista dos avaliados por ordem.
        lista_avaliados = []
        for valor, avaliado in avaliação[2:-2]:
            if avaliado not in lista_avaliados:
                lista_avaliados.append(avaliado)

                # Seguindo esta mesma ordem, completar a lista auxiliar com os valores.
        for avaliado1 in lista_avaliados:

            lista_aux = []
            lista_aux.append(avaliação[1])
            lista_aux.append(avaliado1)

            for valor, avaliado2 in avaliação[2:-2]:
                if avaliado1 == avaliado2:
                    lista_aux.append(valor)

            lista_aux.append(avaliação[-1])
            lista_aux.append(avaliação[-2])

            lista_final.append(lista_aux)

    for i in range(len(lista_final)):

        k1 = 1
        k2 = 1
        obs1 = lista_final[i][-1]
        obs2 = lista_final[i][-2]

        if obs1 == "":
            k1 = 0
        if obs2 == "":
            k2 = 0

        for j in range(i + 1, len(lista_final)):
            if lista_final[j][-1] == obs1 and k1 == 1:
                lista_final[j][-1] = ""
            if lista_final[j][-2] == obs2 and k2 == 1:
                lista_final[j][-2] = ""

                # Output
    wb2 = xl.load_workbook(caminho_FINAL)
    ws2 = wb2.active

    sheet = wb2["S2.1"]
    for i in range(1, len(lista_final) + 1):
        for j in range(1, len(lista_final[0]) + 1):
            sheet.cell(row=i + 2, column=j).value = lista_final[i - 1][j - 1]

    wb2.save(caminho_FINAL)

    progress_step(13)
    root.update_idletasks()

    #######################################################################################################################################
    """
    S2.2 - Import de excel necessário e escrita no ficheiro final
    """

    base_dados_ex = pd.read_excel(caminho_S22)
    base_dados = base_dados_ex.values.tolist()

    # Recolha de dados

    # Cruzamento entre PM's e projetos para as duas primeiras colunas

    colunas = []
    for coluna in base_dados_ex.columns:
        colunas.append(coluna)

    # lista_projeto_pm=[["Responsabilidade Social (Interno)","Pedro Machado"],["Site EPIC","Liliana Freitas"],["WEB4SOCIETY","Pedro Machado"],["RefractorAJ","Filipe AraÃƒÂƒÃ‚Âºjo"]]

    base_dados2 = list(base_dados)

    # Limpar timestamps
    for i in range(len(base_dados2)):
        del base_dados2[i][0]

    # Puxar dados para a esquerda
    for i in range(len(base_dados2)):
        k = 0
        for j in range(len(colunas) - 1):
            valor = base_dados2[i][j + k]
            try:
                if np.isnan(valor) == True:
                    del base_dados2[i][j + k]
                    k -= 1
            except:
                continue

    # Completar ÃƒÂƒÃ‚Âºltima coluna opcional.
    for i in range(len(base_dados2)):
        if len(base_dados2[i]) != 17:
            base_dados2[i].append("(sem comentários)")

    # Lista de projetos
    lista_projetos = []
    for linha in base_dados2:
        if linha[0] not in lista_projetos:
            lista_projetos.append(linha[0])

    # Construção de linhas
    final = []
    for projeto, pm in lista_projeto_pm:
        for linha in base_dados2:
            aux = []
            if linha[0] == projeto:
                aux.append(linha[1])
                aux.append(pm)
                for x in linha[2:]:
                    aux.append(x)
                final.append(aux)

    # Output
    wb2 = xl.load_workbook(caminho_FINAL)
    ws2 = wb2.active

    sheet = wb2["S2.2"]
    for i in range(1, len(final) + 1):
        for j in range(1, len(final[0]) + 1):
            sheet.cell(row=i + 2, column=j).value = final[i - 1][j - 1]

    wb2.save(caminho_FINAL)

    progress_step(13)
    root.update_idletasks()

    #######################################################################################################################################
    """
    S2.3 - Import de excel necessário e escrita no ficheiro final
    """
    
    # LEITURA DA BASE DE DADOS

    base_dados_ex = pd.read_excel(caminho_S23)
    base_dados = base_dados_ex.values.tolist()

    colunas = []
    for coluna in base_dados_ex.columns:
        colunas.append(coluna)
    colunas[2] = "Avaliado"
    del colunas[0]

    # EXTRAÇÃO DE INFORMAÇÃO

    # Avaliadores
    lista_avaliadores = []
    for i in range(len(base_dados)):
        if base_dados[i][1] not in lista_avaliadores:
            lista_avaliadores.append(base_dados[i][1])

    # Projetos
    lista_projetos = []
    for i in range(len(base_dados)):
        if base_dados[i][2] not in lista_projetos:
            lista_projetos.append(base_dados[i][2])

    for i in range(len(lista_projetos)):
        aux = ""
        for j in range(len(lista_projetos[i]) - 1):
            if lista_projetos[i][j + 1] != "(":
                aux += lista_projetos[i][j]
            if lista_projetos[i][j] == "(":
                break
        lista_projetos[i] = aux[:-1]

        # Coordenadores
    lista_coordenadores = []
    for i in range(len(base_dados)):
        k = 0
        aux1 = ""
        for letra in base_dados[i][2]:
            if letra == "(":
                k = 1
            elif k == 1:
                aux1 += letra
            else:
                continue
        k = 0
        aux2 = ""
        for letra in aux1:
            if letra.isupper() == True:
                k = 1
            if k == 1:
                aux2 += letra
            else:
                continue
        lista_coordenadores.append(aux2[:-1])
    lista_coordenadores = list(set(lista_coordenadores))

    # CONSTRUIR O EXCEL

    # 1ª e 2ª coluna
    lista_coluna_1e2 = []
    for projeto in lista_projetos:
        # append coordenador
        for i in range(len(base_dados)):
            if projeto in base_dados[i][2]:
                if base_dados[i][1] in lista_coordenadores:
                    coordenador = base_dados[i][1]
                else:
                    pm = base_dados[i][1]
        lista_coluna_1e2.append([projeto, coordenador, coordenador])
        lista_coluna_1e2.append([projeto, pm, coordenador])

    # Com base nas chaves primárias e secundárias já apontadas nas primeiras colunas, completar.

    final = []
    for projeto, avaliador, avaliado in lista_coluna_1e2:
        for i in range(len(base_dados)):
            if projeto in base_dados[i][2] and avaliador in base_dados[i][1]:

                aux = []
                aux.append(avaliador)
                aux.append(avaliado)
                for j in range(3, len(base_dados[i])):
                    aux.append(base_dados[i][j])

                final.append(aux)

    # Passar os valores para a folha 2.3 do excel final.

    # Output
    wb2 = xl.load_workbook(caminho_FINAL)
    ws2 = wb2.active

    sheet = wb2["S2.3"]
    for i in range(1, len(final) + 1):
        for j in range(1, len(final[0]) + 1):
            sheet.cell(row=i + 2, column=j).value = final[i - 1][j - 1]

    wb2.save(caminho_FINAL)

    progress_step(13)
    root.update_idletasks()
    
    #######################################################################################################################################
    """
    S1 - XXX
    """

    wb1 = xl.load_workbook(caminho_ADS1)
    ws1 = wb1.worksheets[0]

    # DESTINO
    wb2 = xl.load_workbook(caminho_FINAL)
    ws2 = wb2.active
    base_dados_ex = pd.read_excel(caminho_ADS1)
    base_dados = base_dados_ex.values.tolist()
    #códido para retirar os avaliadores e avaliados
    lista_avaliadores = []
    for i in range(len(base_dados)):
        if base_dados[i][1] not in lista_avaliadores:
            lista_avaliadores.append(base_dados[i][1])
    lista_avaliados = lista_avaliadores.copy()
    lista_avaliados.sort()

    sheet = wb2["AD S1"]
    check = 0
    col = 0

    # Eliminar diretor (o diretor não é avaliado mas avalia)
    if diretor.__len__() >= 2:
        lista_avaliados.remove(diretor[0])
        lista_avaliados.remove(diretor[1])
    else:
        lista_avaliados.remove(diretor[0])
    #colocar o avaliador e os respetivos avaliados no excel
    for y in range(lista_avaliadores.__len__()):
        sheet.cell(row=(y+3)+y*(lista_avaliados.__len__()-1), column=1).value = lista_avaliadores[y]
        for n in range(lista_avaliados.__len__()):
            sheet.cell(row=(n+3)+y*(lista_avaliados.__len__()),column=2).value=lista_avaliados[n]

    for y in range(lista_avaliadores.__len__()):
        for x in range(base_dados_ex.columns.__len__()-2):
            c = ws1.cell(row=2+y, column=x + 3)
            if check != (lista_avaliadores.__len__()-2):
                check = check + 1
                sheet.cell(row=check+2+y*(lista_avaliados.__len__()), column=3+col).value = c.value
            else:
                col=col+1
                check=1
                sheet.cell(row=check+2+y*(lista_avaliados.__len__()), column=3 + col).value = c.value
        col=0
        check=0

    wb2.save(caminho_FINAL)

    # PROGRESS
    progress_step(13)
    root.update_idletasks()

    #######################################################################################################################################
    """
    M1 - XXX
    """
    
    #Lista a averiguar (por causa do problema com a concordância com S1) -> cuidado no executável.
    #lista_avaliados=['Ana Isabel Martins', 'Daniel Cunha', 'Francisco Salgado', 'Gonçalo Arantes', 'Hélder Lobo', 'José Soares', 'Pedro Vieira', 'Sofia Gonçalves', 'Sónia Fernandes', 'Tomás Silva']
    
    wb1 = xl.load_workbook(caminho_ADM1)
    ws1 = wb1.worksheets[0]
    # DESTINO
    wb2 = xl.load_workbook(caminho_FINAL)
    ws2 = wb2.active

    base_dados_ex = pd.read_excel(caminho_ADM1)
    base_dados = base_dados_ex.values.tolist()
    sheet = wb2["AD M1"]
    check = 0
    col = 0
    aux=0
    for y in range(3 * diretor.__len__()):  # código feito para 3 meses
        sheet.cell(row=((y + 3 + aux) + y * (lista_avaliados.__len__() - 1)), column=1).value = ws1.cell(
            row=(base_dados_ex.__len__() + diretor.__len__() - 3 * diretor.__len__()) + y, column=2).value # vai buscar o avalaidor ao excel
        for n in range(lista_avaliados.__len__()):
            sheet.cell(row=((n + 3 + aux) + y * (lista_avaliados.__len__())), column=2).value = lista_avaliados[n] # coloca os avaliados na posição correto do excel
    numero_perguntas = 0
  
    for y in range(3 * diretor.__len__()):  # numero de meses
        for x in range(base_dados_ex.columns.__len__() - 2):
            c = ws1.cell(row=(base_dados_ex.__len__() + 2 - 3 * diretor.__len__()) + y, column=x + 3)
            if numero_perguntas < int((base_dados_ex.columns.__len__() - 4) / lista_avaliados.__len__()): # numero de perguntas menos as 2 últimas
                if check != (lista_avaliados.__len__()):
                    check = check + 1
                    sheet.cell(row=check + 2 + y * (lista_avaliados.__len__()), column=3 + col).value = c.value
                else:
                    col = col + 1
                    check = 1
                    numero_perguntas = numero_perguntas + 1
                    sheet.cell(row=check + 2 + y * (lista_avaliados.__len__()), column=3 + col).value = c.value
            else: #código especifico para as duas últimas questões
                col = col + 1
                sheet.cell(row=y * (lista_avaliados.__len__()) + 3, column=3 + col).value = c.value
        numero_perguntas = 0
        col = 0
        check = 0
        
    wb2.save(caminho_FINAL)

    # PROGRESS
    progress_step(9)
    root.update_idletasks()

    # Label de conclusão
    Label(root, text="CONGRATS, YOU HAVE BEEN RESCUED", foreground='Green', font=tipo_letra_negrito).place(x=430, y=625)


# Botão de saÃ­da
QUITbtn = tk.Button(root, text="Tenho que ir", font=tipo_letra_negrito, command=lambda: root.destroy()).place(x=1100,y=640)


###
def help_f():
    url = "https://drive.google.com/drive/u/2/folders/1PcWYlCabkD13CQqQaUfOBDqS5TRjjQGT"
    webbrowser.open(url, new=0, autoraise=True)


# Botão de ajuda
HELPbtn = tk.Button(root, text="Ajuda-me pf", font=tipo_letra_negrito, command=lambda: help_f()).place(x=10, y=640)

#######################################################################################################################################

# Botão executar
EXECbtn = tk.Button(root, width=10, height=5, text="EXECUTAR", font=tipo_letra_negrito, bg="#FF914D",
                    command=lambda: EXEC_fun(caminho_FINAL, caminho_ADM1, caminho_ADS1, caminho_AM1, caminho_AM2,
                                             caminho_AM3, caminho_S21, caminho_S22, caminho_S23, departamento)).place(x=950, y=400)

# Final
root.mainloop()